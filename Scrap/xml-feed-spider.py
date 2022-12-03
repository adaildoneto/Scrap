import pprint
from bs4 import BeautifulSoup as bs
import requests
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import pandas as pd

# criando a planilha workbook
wb = Workbook()

# ativando a planilha workbook
ws1 = wb.active

# Renomeando a planilha workbook geral
ws1.title = 'Olho de Sauron'

#criando a primeira linha da tabela para servir de guia
pdata = ['Data', 'Hora', 'Titulo', 'Descricao', 'Url',]
ws1.append(pdata)


sites = ['http://ac24horas.com', 'http://contilnetnoticias.com.br', 'http://correio68.com', 'https://agencia.ac.gov.br', 'https://nahoradanoticia.com.br',
'http://folhadoacre.com.br', 'http://yaconews.com', 'http://jornalopiniao.net', 'http://3dejulhonoticias.com.br', 'https://noticiadoacre.com.br',
'http://ecosdanoticia.net.br', 'http://agazetadoacre.com', 'http://www.acre.com.br', 'https://acreinfoco.com', 'https://acjornal.com',
'http://oaltoacre.com', 'http://agazeta.net', 'http://noticiasdahora.com.br',  'http://acreagora.com', 'https://www.noticiasdafronteira.com.br', 
'https://www.juruaonline.com.br', 'https://www.juruaemtempo.com.br',]

jsite = len (sites)

for i in range (jsite):
    url = (sites[i] + '/feed')
    response = requests.get(url)
    soup = bs (response.content, features='xml')
    items = soup.find_all('item')
    print(url)
    for item in items :
        titulo = item.find('title').text
        data1 = item.find('pubDate').text
        data = data1[4:16] 
        hora = data1[17:25]
        #autor = item.find('dc:creator').text
        descricao = item.find('description').text
        url = item.find('link').text
        img = item.find('img')
        #conteudo = item.find('content:encoded').text

        #estruturando o conteudo dentro da celula
        pdata = (data, hora, titulo, descricao, url)
        
           
        # ativando a planilha workbook
        ws1.append(pdata)

# Salvando o planilha
wb.save('olhodesauron.xlsx')


# lendo a planilha The Eye of Thundera
df1 = pd.read_excel('olhodesauron.xlsx')

# ordenando a planilha por data e hora
ok = df1.sort_values(by=['Data', 'Hora'])
print ('Planinha Atualizada')

ok.to_json(r'sauron-ordenado.json')
print ('Criando Json')

ok.to_excel("sauron-ordenado.xlsx")
print ('Planinha Atualizada')
