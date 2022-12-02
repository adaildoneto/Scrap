
from bs4 import BeautifulSoup as bs
import requests
import json
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import pandas as pd

# criando a planilha workbook
wb = Workbook()

# ativando a planilha workbook
ws1 = wb.active

# Renomeando a planilha workbook geral
ws1.title = 'Olho de Thundera'

#criando a primeira linha da tabela para servir de guia
pdata = ['Data', 'Hora','Titulo', 'Conteúdo', 'Url', 'Imagem']
ws1.append(pdata)

#criando a planilha Polícia
ws2 = wb.create_sheet(title="Sem Data")
ws2.append(pdata)



sites = ['http://ac24horas.com', 'http://contilnetnoticias.com.br', 'http://correio68.com', 'https://agencia.ac.gov.br',
'http://folhadoacre.com.br', 'http://yaconews.com', 'http://jornalopiniao.net', 'http://3dejulhonoticias.com.br', 'https://www.noticiasdahora.com.br',
'http://ecosdanoticia.net.br', 'http://agazetadoacre.com', 'http://www.acre.com.br', 'http://acreagora.com', 'https://acreinfoco.com',
'http://oaltoacre.com', 'http://agazeta.net', 'http://noticiasdahora.com.br',  'http://oacreagora.com', ]

jnews = len (sites)

for i in range (jnews):
    turl = (sites[i] + '/wp-json/wp/v2/search/?subtype=post&search=Cameli')
    response = requests.get(turl)

    data = response.text
    
    if response.status_code == 200:  
      print('site ' + turl + ' ok!')
      dados = json.loads(data)
      for j in dados:
         titulo = j['title']
         link = j['url']
         
         id_ = str(j['id'])
        
         pturl = (sites[i] + '/wp-json/wp/v2/posts/' + id_)
         response = requests.get(pturl)
         artigo = response.text
         data_json = json.loads(artigo)
         k = data_json
         try:
            cdata = k['date']
            cdata1 = cdata[0:10]
            chora = cdata[11:20]
         except KeyError:
            cdata = 0
         try:
            content = k['content']['rendered']
         except KeyError:
            content = 'Não encontrado'
         try:
             img = k['jetpack_featured_media_url']
         except KeyError:
             turl = (sites[i] + '/wp-json/wp/v2/media/' + id_)
             response = requests.get(turl)
             if response.status_code == 200:
                  imagem = response.text
                  i_json = json.loads(imagem)
                  print (i_json)
                  img = i_json['yoast_head_json']['og_image']['url']
             else:
                  img = 'Não encontrado'
         
         
         #estruturando o conteudo dentro da celula
         pdata = (cdata1, chora, titulo, content, link, img )

         # Criando uma nova planilha para a palavra chave Bocalom
         if cdata == 0:
            active_sheet = wb['Sem Data']
            ws2.append(pdata)
         else:
         # ativando a planilha workbook
            active_sheet = wb['Olho de Thundera']
            ws1.append(pdata)
    else:
      print('site ' + turl + ' não habilitado para o json')

# Salvando o planilha
print ('Salvando a planilha')
wb.save('theeye.xlsx')

# lendo a planilha The Eye of Thundera
df1 = pd.read_excel('theeye.xlsx')

# ordenando a planilha por data e hora
ok = df1.sort_values(by=['Data', 'Hora'])
print ('Planinha Atualizada')

ok.to_json(r'ordenado.json')
print ('Criando Json')

ok.to_excel("ordenado.xlsx")
print ('Planinha Atualizada')