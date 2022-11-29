from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter
import requests

# criando a planilha workbook
wb = Workbook()

# ativando a planilha workbook
ws1 = wb.active

# Renomeando a planilha workbook geral
ws1.title = 'Ac24horas Scraping'

#criando a primeira linha da tabela para servir de guia
pdata = ['Data', 'Link da Imagem', 'Categoria', 'Titulo', 'Link', 'Texto Completo', 'Descricao', 'Autor', ]
ws1.append(pdata)

#criando a planilha cameli
ws2 = wb.create_sheet(title="Cameli")
pdata = ['Data', 'Link da Imagem', 'Categoria', 'Titulo', 'Link', 'Texto Completo', 'Descricao', 'Autor', ]
ws2.append(pdata)

#criando a planilha Bocalom 
ws3 = wb.create_sheet(title="Bocalom")
pdata = ['Data', 'Link da Imagem', 'Categoria', 'Titulo', 'Link', 'Texto Completo', 'Descricao', 'Autor', ]
ws3.append(pdata)

for i in range(1,31): 
    #configurando a paginação da url
    if i == 1:
        url = "https://ac24horas.com/ultimas-noticias/"
        response  = requests.get(url)
    else:
        purl = "https://ac24horas.com/ultimas-noticias/page/"
        page = i - 1
        url = purl + str(page)
        response  = requests.get(url)
    
    #parse da url principal e paginação
    soup = bs (response.content, 'lxml')
    title = soup.find_all('li', class_="mvp-blog-story-wrap")
    
    #extração dos dados gerais
    for t in title:
        Categoria = t.find(class_="mvp-cd-cat").get_text(strip=True)
        
        #Caso de exceção, se por acaso nao tiver imagem destacada
        try:
            img = t.find('img').get('src')
        except AttributeError:
            img = 'Imagem não encontrada'
            
        Titulo = t.find('h2').get_text(strip=True)
        Link = t.find('a').get('href')

        #Extração do conteudo de cada matéria   
        conteudoUrl = t.find('a').get('href')
        contenido = requests.get(conteudoUrl)
        conteudoex = bs (contenido.content, 'lxml')

        #caso de exceção, se por acaso o nome do autor for omitido          
        try:
            Autor = conteudoex.find('span', class_="author-name").get_text(strip=True)
        except AttributeError:
            Autor = 'Autor não encontrado'

        try:        
            Data = conteudoex.find('time', class_="post-date").get_text(strip=True)
        except AttributeError:
            Data = 'Data não encontrada'  

        try: 
            Textocompleto = conteudoex.find("div", id="mvp-content-main").get_text(strip=True)
        except AttributeError:
            Textocompleto = 'Texto não encontrada'  

        Descricao = t.find('p').get_text(strip=True)

        #criando a estrutura do dados       
        pdata = [Data, img, Categoria, Titulo, Link, Textocompleto, Descricao, Autor]
        #Adicionando o conteudo a planilha geral
        #Criando uma nova planilha para a palavra chave Cameli
        if 'Cameli' in Textocompleto:
           active_sheet = wb['Cameli']
           ws2.append(pdata)

        # Criando uma nova planilha para a palavra chave Bocalom
        if 'Bocalom' in Textocompleto:
           active_sheet = wb['Bocalom']
           ws3.append(pdata)
    
        # ativando a planilha workbook
        
        ws1.append(pdata)

       
# Salvando o planilha
wb.save('spider.xlsx')